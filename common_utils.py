import os
import calendar
import pyodbc
from openpyxl.styles import Border, Side, Font, PatternFill
from PySide6.QtGui import QIcon
import config

def center_window(win, width, height):
    """ウィンドウを画面中央に配置し、サイズを固定する共通関数"""
    win.update_idletasks()
    screen_width = win.winfo_screenwidth()
    screen_height = win.winfo_screenheight()
    
    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))
    
    win.geometry(f"{width}x{height}+{x}+{y}")

def set_common_window_icon(window_obj, icon_name="my_logo.ico"):
    """
    指定されたウィンドウオブジェクトに共通のアイコンを設定する。
    """
    try:
        # assets フォルダ内のアイコンファイルを指定
        current_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(current_dir, "assets", icon_name)
        
        if os.path.exists(icon_path):
            window_obj.setWindowIcon(QIcon(icon_path))
    except Exception as e:
        print(f"Icon error: {e}")

def get_date_info(year_val, month_val):
    """入力値から日付範囲、シート名、保存先を生成する"""
    try:
        y = int(year_val)
        m = int(month_val)
        m_str = str(m).zfill(2)
        _, last_day = calendar.monthrange(y, m)
        
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        
        return {
            "start": f"{y}-{m_str}-01",
            "end": f"{y}-{m_str}-{last_day}",
            "sheet_name": f"{y}年{m}月",
            "desktop_path": desktop,
            "year": y,
            "month": m
        }
    except ValueError:
        return None

def get_excel_styles():
    """共通のExcelスタイル一式を返す"""
    side = Side(style='thin', color='000000')
    return {
        "border": Border(top=side, bottom=side, left=side, right=side),
        "font": Font(name=config.FONT_NAME),
        "font_bold": Font(name=config.FONT_NAME, bold=True),
        "fill_header": PatternFill(start_color=config.BG_HEADER, end_color=config.BG_HEADER, fill_type='solid'),
        "fill_total": PatternFill(start_color=config.BG_TOTAL, end_color=config.BG_TOTAL, fill_type='solid')
    }

def get_db_connection():
    """データベースに接続する"""
    return pyodbc.connect(config.CONN_STR)

def handle_window_close(current_window, parent_menu=None, parent_root=None):
    """【共通】子画面が閉じる際に親画面（メニュー）を安全に再表示する共通処理"""
    if parent_menu:
        if hasattr(parent_menu, "show"):
            parent_menu.show()
        elif hasattr(parent_menu, "show_menu"):
            parent_menu.show_menu()
        elif hasattr(parent_menu, "deiconify"):
            parent_menu.deiconify()
    elif parent_root:
        if hasattr(parent_root, "deiconify"):
            parent_root.deiconify()
        if hasattr(parent_root, "lift"):
            parent_root.lift()
    current_window.close()


def disable_dummy_buttons_tab_focus(window_ui, prefix="btn_dummy_"):
    """【共通】特定の接頭辞を持つダミーボタンを見つけた場合、Tabキーでの選択のみをスキップさせる（画面上には表示したまま）"""
    from PySide6.QtWidgets import QPushButton
    from PySide6.QtCore import Qt
    
    all_buttons = window_ui.findChildren(QPushButton)
    for btn in all_buttons:
        if btn.objectName().startswith(prefix):
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)


def setup_text_edits_tab_focus(window_obj):
    """【共通】画面内のすべての QTextEdit で Tab キーによるフォーカス移動を有効化する"""
    from PySide6.QtWidgets import QTextEdit
    all_edits = window_obj.findChildren(QTextEdit)
    for edit in all_edits:
        edit.setTabChangesFocus(True)

