import calendar
import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QDate, Qt
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox
from ui_files.app_nouhin_ui import Ui_MainWindow

# 共通ユーティリティのインポート
import common_utils

# Pythonに古い一時ファイル（.pyc）を作らせない設定
sys.dont_write_bytecode = True

# UIファイルをロード
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)

# --- Excel列位置・集計項目のマジックナンバー定数化 ---
COL_DATE = 1          # 日付列
COL_TOTAL_AMOUNT = 7  # 売上合計列


class NouhinWindow(QMainWindow, Ui_MainWindow):
    """納品書に基づく売上内訳 ウィンドウ管理クラス"""

    def __init__(self, parent_menu=None):
        """初期化処理・コンポーネントとUIのセットアップ"""
        super().__init__()
        self.parent_menu = parent_menu
        self.setupUi(self)
        self.init_ui()

    def init_ui(self):
        """UIの初期セットアップ（外観・シグナル接続）"""
        self.setupUi(self)
        self.setWindowTitle("納品書に基づく売上内訳")
        common_utils.set_common_window_icon(self)

        # 初期値として現在のシステム日付（今日）をセット
        if hasattr(self, "date_target_year_month"):
            self.date_target_year_month.setDate(QDate.currentDate())

        # Tabキーフォーカスの最適化（初期フォーカス）
        if hasattr(self, "date_target_year_month"):
            self.date_target_year_month.setFocus()

        # ★【共通部品化】ダミーボタンの表示は維持し、Tabフォーカスのみを一括無効化
        common_utils.disable_dummy_buttons_tab_focus(self)

        # ★【共通部品化】すべての複数行テキストエリア（QTextEdit）でTabキー移動を自動有効化
        common_utils.setup_text_edits_tab_focus(self)

        # 整理したオブジェクト名でシグナル（イベント）接続
        if hasattr(self, "btn_back"):
            self.btn_back.clicked.connect(self.close_window)
        if hasattr(self, "btn_clear"):
            self.btn_clear.clicked.connect(self.clear_ui)
        if hasattr(self, "btn_exe_output"):
            self.btn_exe_output.clicked.connect(self.run_query)

    def clear_ui(self):
        """【クリアボタン（btn_clear）押下時】入力を画面起動時の状態に戻す"""
        if hasattr(self, "date_target_year_month"):
            self.date_target_year_month.setDate(QDate.currentDate())

    def close_window(self):
        """【共通関数呼び出し】親メニュー画面を安全に最前面表示させて自身を閉じる"""
        common_utils.handle_window_close(self, self.parent_menu)

    def closeEvent(self, event):
        """【×ボタン押下時】戻るボタンと同じ終了処理をフックする"""
        self.close_window()
        event.accept()

    def run_query(self):
        """【出力ボタン（btn_exe_output）押下時】条件を取得しSQLを実行、Excelに出力する"""
        if hasattr(self, "date_target_year_month"):
            q_date = self.date_target_year_month.date()
            year_val = q_date.year()
            month_val = q_date.month()
        else:
            QMessageBox.critical(self, "システムエラー", "日付ボックスが見つかりません。")
            return

        start_date = f"{year_val:04d}-{month_val:02d}-01"
        _, last_day = calendar.monthrange(year_val, month_val)
        end_date = f"{year_val:04d}-{month_val:02d}-{last_day:02d}"
        sheet_name = f"{year_val:04d} 年{month_val:02d} 月"
        desktop_path = os.path.expanduser("~/Desktop")
        file_name = f"納品書に基づく売上内訳_{year_val} 年{month_val} 月.xlsx"
        save_path = os.path.join(desktop_path, file_name)
        
        conn = common_utils.get_db_connection()

        # try...finally による安全な切断管理
        try:
            # SQLクエリ文（ご指示通り完全未変更）
            query = """
            SELECT URH_DENDAT AS 日付
            ,SUM(CASE WHEN SHO_KBN = 3 AND URM_SHOCD NOT IN ('806000', '806005', '807000') THEN URM_URIKIN ELSE 0 END) AS 商品
            ,SUM(CASE WHEN URM_SHOCD = '806005' THEN URM_URIKIN ELSE 0 END) AS 荷造梱包費
            ,SUM(CASE WHEN URM_SHOCD = '807000' THEN URM_URIKIN ELSE 0 END) AS 雑収入
            ,SUM(CASE WHEN URM_SHOCD = '806000' THEN URM_URIKIN ELSE 0 END) AS 運賃
            ,SUM(CASE WHEN SHO_KBN = 2 THEN URM_URIKIN ELSE 0 END) AS 製品
            ,0 AS 売上合計
            ,SUM(CASE WHEN URM_SHOCD = '805000' THEN URM_SURYO ELSE 0 END) AS TPH_数量
            ,SUM(CASE WHEN URM_SHOCD = '805000' THEN URM_URIKIN ELSE 0 END) AS TPH_金額
            ,SUM(CASE WHEN URM_SHOCD = '804001' THEN URM_SURYO ELSE 0 END) AS 電纜_数量
            ,SUM(CASE WHEN URM_SHOCD = '804001' THEN URM_URIKIN ELSE 0 END) AS 電纜_金額
            FROM T_URHDAT
            INNER JOIN T_URMDAT ON URH_KCODE = URM_KCODE AND URH_DENNO = URM_DENNO
            LEFT JOIN T_SHOMST ON URM_SHOCD = SHO_SHOCD
            WHERE URH_DENDAT BETWEEN ? AND ?
            GROUP BY URH_DENDAT
            ORDER BY URH_DENDAT
            """
            df = pd.read_sql(query, conn, params=[str(start_date), str(end_date)])
            
            if df.empty:
                QMessageBox.information(self, "結果", "該当するデータは見つかりませんでした。")
                return

            # Excelへの書き込み
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
            wb = load_workbook(save_path)
            ws = wb.active
            s = common_utils.get_excel_styles()
            from openpyxl.styles import Alignment
            from copy import copy
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 元の通り、2段目のヘッダー文字列を個別にセット
            ws['H2'] = "数量"
            ws['I2'] = "金額"
            ws['J2'] = "数量"
            ws['K2'] = "金額"

            # 【エラー修正】元の安定した2重ループによるセルアクセス構造に完全復元
            for i in range(1, ws.max_column + 1):
                col_letter = get_column_letter(i)
                for r in range(1, ws.max_row + 1):
                    cell = ws.cell(row=r, column=i)
                    cell.border = s["border"]

                    # 1〜2行目（ヘッダー行）の装飾
                    if cell.row <= 2:
                        cell.fill = s["fill_header"]
                        cell.alignment = center_align
                        if cell.coordinate == "D1":
                            small_font = copy(s["font_bold"])
                            small_font.sz = 9
                            cell.font = small_font
                        else:
                            cell.font = s["font_bold"]
                    else:
                        # 明細行のフォント設定
                        cell.font = s["font"]

                    # 3行目以降（データ行）の数値フォーマットと数式
                    if cell.row >= 3:
                        if i == COL_DATE:
                            cell.number_format = 'm"月"d"日"'
                        elif i >= 2:
                            cell.number_format = '#,##0'

                        if i == COL_TOTAL_AMOUNT:
                            cell.value = f"=SUM(B{cell.row}:F{cell.row})"

                # 列幅の最適化
                if i == 1:
                    ws.column_dimensions[col_letter].width = 10
                elif i in (3, 5):
                    ws.column_dimensions[col_letter].width = 8
                elif i in (8, 10):
                    ws.column_dimensions[col_letter].width = 10
                else:
                    ws.column_dimensions[col_letter].width = 15

            # すべての値を配置し終わった「最後」に、セルを結合する
            header_map = [
                ('A1:A2', '日付'), ('B1:B2', '商品売上'),
                ('C1:C2', '荷造梱包費'), ('D1:D2', '雑収入\n（梱包破損補償等）'),
                ('E1:E2', '運賃'), ('F1:F2', '製品'), ('G1:G2', '売上合計'),
                ('H1:I1', 'TPH8512RED'), ('J1:K1', '電纜ホース')
            ]
            for cell_range, title in header_map:
                ws.merge_cells(cell_range)
                
                # 結合の左上セルにタイトルを設定
                for row in ws[cell_range]:
                    for cell in row:
                        cell.value = title
                        break
                    break

            # 最終行（合計行）の追加と装飾
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1).value = "合計"
            for c in range(2, ws.max_column + 1):
                col_let = get_column_letter(c)
                target_cell = ws.cell(row=last_row, column=c)
                target_cell.value = f"=SUM({col_let}3:{col_let}{last_row-1})"
                target_cell.number_format = '#,##0'

            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=last_row, column=c)
                cell.border = s["border"]
                cell.fill = s["fill_total"]
                cell.font = s["font_bold"]

            ws.freeze_panes = 'A3'
            wb.save(save_path)
            QMessageBox.information(self, "完了", f"出力が完了しました。\n 保存先: {save_path}")

        except PermissionError:
            QMessageBox.critical(self, "エラー", "Excelファイルが開いています。閉じてから再実行してください。")
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"処理中にエラーが発生しました:\n{e}")
        finally:
            # 例外発生時も確実にDBを切断する堅牢設計
            conn.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = NouhinWindow()
    window.show()
    sys.exit(app.exec())
