import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QTextCursor
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QPushButton
from ui_files.app_urikake_ui import Ui_MainWindow
from ui_files.app_urikake_setting_ui import Ui_SubWindow

# 共通ユーティリティのインポート
import common_utils

# Pythonに古い一時ファイル（.pyc）を作らせない設定
sys.dont_write_bytecode = True

# --- Excel列位置・集計項目のマジックナンバー定数化 ---
COL_CODE = 2    # コード列
COL_AMOUNT = 5  # 売上金額列

class UrikakeWindow(QMainWindow, Ui_MainWindow):
    """売掛金回収状況一覧 メイン画面クラス"""

    def __init__(self, parent_menu=None):
        """初期化処理・コンポーネントのセットアップ"""
        super().__init__()
        self.parent_menu = parent_menu
        self.init_ui()

    def init_ui(self):
        """UIの初期セットアップとシグナル接続"""
        self.setupUi(self)
        common_utils.set_common_window_icon(self)

        # 画面起動時の状態をセット (現在の年月)
        self.reset_to_initial_state()

        # Tabキーフォーカスの最適化（初期フォーカス）
        if hasattr(self, "date_target_year_month"):
            self.date_target_year_month.setFocus()

        # ★【共通部品化】ダミーボタンの表示は維持し、Tabフォーカスのみを一括無効化
        common_utils.disable_dummy_buttons_tab_focus(self)

        # ★【共通部品化】すべての複数行テキストエリア（QTextEdit）でTabキー移動を自動有効化
        common_utils.setup_text_edits_tab_focus(self)

        # 整理したオブジェクト名でシグナル（イベント）接続
        if hasattr(self, "btn_back"):
            self.btn_back.clicked.connect(self.close_window)
        if hasattr(self, "btn_exe_setting"):
            self.btn_exe_setting.clicked.connect(self.show_setting_window)
        if hasattr(self, "btn_clear"):
            self.btn_clear.clicked.connect(self.clear_fields)
        if hasattr(self, "btn_exe_output"):
            self.btn_exe_output.clicked.connect(self.run_query)

    def reset_to_initial_state(self):
        """日付ボックスを画面起動時の状態（現在の年月）に戻す"""
        now = datetime.now()
        if hasattr(self, "date_target_year_month"):
            self.date_target_year_month.setDate(QDate(now.year, now.month, 1))

    def clear_fields(self):
        """クリアボタン 押下時：画面起動時の状態に戻す"""
        self.reset_to_initial_state()

    def close_window(self):
        """【共通関数呼び出し】親メニュー画面を安全に最前面表示させて自身を閉じる"""
        common_utils.handle_window_close(self, self.parent_menu)

    def closeEvent(self, event):
        """×ボタンクリック時のイベント制御（メニュー再表示）"""
        self.close_window()
        event.accept()

    def show_setting_window(self):
        """設定変更サブ画面の呼び出し"""
        self.setting_win = SettingWindow(self)
        self.setting_win.show()

    def run_query(self):
        """出力ボタン (btn_exe_output) 押下時のExcel出力メインロジック"""
        if hasattr(self, "date_target_year_month"):
            qdate = self.date_target_year_month.date()
            y_in = str(qdate.year())
            m_in = str(qdate.month())
        else:
            QMessageBox.critical(self, "システムエラー", "日付ボックス(date_target_year_month)が見つかりません。")
            return

        dt = common_utils.get_date_info(y_in, m_in)
        if not dt:
            QMessageBox.warning(self, "入力エラー", "対象年月を正しく選択してください")
            return

        conn = common_utils.get_db_connection()
        cursor = conn.cursor()
        file_name = f"売掛金回収状況一覧_{dt['year']} 年{dt['month']} 月.xlsx"
        save_path = f"{dt['desktop_path']}/{file_name}"

        # try...finally による安全な切断管理
        try:
            # --- 1. 新規得意先の同期処理 ---
            sync_query = """
            INSERT INTO Table_2 (code, sort, flag)
            SELECT TOK_TOKCD, 9999, 1 FROM T_TOKMST
            WHERE NOT EXISTS (SELECT 1 FROM Table_2 WHERE Table_2.code = T_TOKMST.TOK_TOKCD)
            """
            cursor.execute(sync_query)
            conn.commit()

            # --- 2. メインのデータ抽出 ---
            query = """
            SELECT TOK_SIMEBI AS 締日, CAST(code AS INT) AS コード, TOK_TOKNM1 AS 得意先名,
            NULL AS 区分, ISNULL(SEK_URIAGE, 0) + ISNULL(SEK_TAX, 0) AS 売上金額,
            NULL AS 入金額, NULL AS 備考, sort AS _sort_val
            FROM Table_2
            LEFT JOIN T_TOKMST ON code = TOK_TOKCD
            LEFT JOIN (
                SELECT SEK_SCODE, SEK_URIAGE, SEK_TAX, SES_SIMEDAT FROM T_SESDAT
                LEFT JOIN T_SEKDAT ON SES_KCODE = SEK_KCODE AND SES_SIMENO = SEK_SIMENO
                LEFT JOIN T_TOKMST ON SEK_KCODE = TOK_KCODE AND SEK_SCODE = TOK_TOKCD
                WHERE SES_SIMEDAT BETWEEN ? AND ?
            ) AS a ON code = SEK_SCODE
            WHERE (Table_2.flag = 1) OR (Table_2.flag = 0 AND (ISNULL(SEK_URIAGE, 0) + ISNULL(SEK_TAX, 0) <> 0))
            ORDER BY sort, code
            """
            df = pd.read_sql(query, conn, params=(dt["start"], dt["end"]))

            if df.empty:
                QMessageBox.information(self, "結果", "該当データはありませんでした。")
                return

            # Excel書き込み
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=dt["sheet_name"])

            wb = load_workbook(save_path)
            ws = wb.active
            
            # スタイル定義共通管理方式
            s = common_utils.get_excel_styles()
            sort_col_idx = ws.max_column

            # セル書式・スタイルの適用（安定したインデックス2重ループ構造）
            for r in range(1, ws.max_row + 1):
                is_new = (ws.cell(row=r, column=sort_col_idx).value == 9999)
                for i in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=i)
                    cell.border = s["border"]
                    cell.font = s["font"]
                    
                    if r == 1:
                        cell.fill = s["fill_header"]
                        cell.font = s["font_bold"]
                    elif is_new:
                        cell.fill = PatternFill(fgColor="E2EFDA", fill_type="solid")

                    # 金額・コード列の3桁カンマ区切り
                    if i in [COL_CODE, COL_AMOUNT] and r > 1:
                        cell.number_format = "#,##0"

            # ソート用の一時列を削除
            ws.delete_cols(sort_col_idx)

            # 列幅の自動調整
            for i in range(1, ws.max_column + 1):
                col_letter = get_column_letter(i)
                if i in [COL_CODE, COL_AMOUNT]:
                    ws.column_dimensions[col_letter].width = 13.5
                else:
                    max_len = max((len(str(ws.cell(row=r, column=i).value or '')) for r in range(1, ws.max_row + 1)), default=0)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 10) * 1.2

            # データ入力規則 (D列: 区分)
            dv = DataValidation(type="list", formula1='"でんさい,振込,相殺"', allow_blank=True)
            dv.add(f"D2:D{ws.max_row}")
            ws.add_data_validation(dv)
            ws.freeze_panes = "A2"

            # 合計行の追加
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=3).value = "合計"
            sum_col = get_column_letter(COL_AMOUNT)
            ws.cell(row=last_row, column=COL_AMOUNT).value = f"=SUM({sum_col}2:{sum_col}{last_row-1})"
            ws.cell(row=last_row, column=COL_AMOUNT).number_format = "#,##0"

            for c in range(1, 8):
                cell = ws.cell(row=last_row, column=c)
                cell.border = s["border"]
                cell.fill = s["fill_total"]
                cell.font = s["font_bold"]

            wb.save(save_path)
            QMessageBox.information(self, "完了", f"出力完了:\n{save_path}")

        except PermissionError:
            QMessageBox.critical(self, "エラー", "Excelを閉じてから実行してください。")
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"失敗しました:\n{e}")
        finally:
            cursor.close()
            conn.close()

class SettingWindow(QMainWindow, Ui_SubWindow):
    """出力設定変更サブ画面 クラス（デザイン維持・挙動完全安定化版）"""

    def __init__(self, parent):
        """初期化処理・ウィンドウの配置とイベント設定"""
        super().__init__(parent)
        self.parent_win = parent
        self.setupUi(self)

        # 子画面として最前面に固定
        self.setWindowModality(Qt.WindowModality.WindowModal)

        # 元の画面（親ウィンドウ）の左上座標を取得し、右下に30ピクセルずつずらして配置
        if self.parent_win:
            parent_geo = self.parent_win.geometry()
            self.move(parent_geo.x() + 30, parent_geo.y() + 30)

        # ★【共通部品化】設定サブ画面内のすべての QTextEdit で Tab フォーカス移動を有効化
        common_utils.setup_text_edits_tab_focus(self)

        # キー入力とフォーカス移動の制御用イベントフィルタを設定
        if hasattr(self, "text_tokcode"):
            self.text_tokcode.installEventFilter(self)
        if hasattr(self, "text_order"):
            self.text_order.installEventFilter(self)

        # Tabキーフォーカス設定
        if hasattr(self, "chk_uriagezero"):
            self.chk_uriagezero.setFocusPolicy(Qt.FocusPolicy.TabFocus)

        # 初期フォーカス設定
        if hasattr(self, "text_tokcode"):
            self.text_tokcode.setFocus()

        # ボタン類のシグナル接続
        self.btn_back.clicked.connect(self.close)
        self.btn_clear.clicked.connect(self.clear_fields)
        self.btn_exe_change.clicked.connect(self.update_settings)

    def eventFilter(self, obj, event):
        """QTextEditでのEnter/Tabキー、およびフォーカスイン・アウトの挙動を完全に制御する"""
        # --- 【追加】フォーカスが当たった（入った）ときの制御 ---
        if event. type() == event. Type. FocusIn:
            # 得意先コード欄、または表示順欄にフォーカスが当たったら全選択（反転）にする
            if obj in (self. text_tokcode, self. text_order):
                # 処理の競合を防ぐため、少しだけ遅延させて全選択を実行します
                from PySide6. QtCore import QTimer
                QTimer. singleShot( 0, obj. selectAll)

        """QTextEditでのEnter/Tabキーおよびフォーカスアウトの挙動を完全に制御する"""
        if event.type() == event.Type.KeyPress:
            # 得意先コード欄での制御
            if obj == self.text_tokcode:
                if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter, Qt.Key.Key_Tab):
                    self.on_enter()
                    if hasattr(self, "text_order"):
                        self.text_order.setFocus()
                        self.text_order.selectAll()
                    return True

            # 表示順欄での制御
            elif obj == self.text_order:
                if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                    self.update_settings()
                    return True
                elif event.key() == Qt.Key.Key_Tab:
                    if hasattr(self, "chk_uriagezero"):
                        self.chk_uriagezero.setFocus()
                    return True

        elif event.type() == event.Type.FocusOut:
            if hasattr(obj, "textCursor"):
                cursor = obj.textCursor()
                if cursor.hasSelection():
                    cursor.clearSelection()
                    obj.setTextCursor(cursor)

        return super().eventFilter(obj, event)

    def on_enter(self):
        """得意先コード入力後の検索・ゼロ埋めと画面表示処理"""
        raw_code = self.text_tokcode.toPlainText().strip()
        if not raw_code:
            return

        code_8 = "".join(filter(str.isdigit, raw_code)).zfill(8)
        self.text_tokcode.blockSignals(True)
        self.text_tokcode.setPlainText(code_8)
        self.text_tokcode.blockSignals(False)

        conn = common_utils.get_db_connection()
        
        # try...finally による安全な切断管理
        try:
            query = "SELECT TOK_TOKNM1, sort, flag FROM Table_2 LEFT JOIN T_TOKMST ON code = TOK_TOKCD WHERE code = ?"
            res = conn.execute(query, (code_8,)).fetchone()
            
            if res:
                cust_name = res[0] if res[0] else ""
                curr_sort = res[1] if res[1] is not None else ""
                curr_flag = res[2]
                self.data_tokname.setText(cust_name)
                if hasattr(self, "text_order"):
                    self.text_order.blockSignals(True)
                    self.text_order.setPlainText(str(curr_sort))
                    self.text_order.blockSignals(False)
                self.chk_uriagezero.setChecked(True if curr_flag == 0 else False)
            else:
                QMessageBox.warning(self, "未登録", "得意先が見つかりません。")
        finally:
            conn.close()

    def clear_fields(self):
        """クリアボタン (btn_clear) 処理"""
        if hasattr(self, "text_tokcode"):
            self.text_tokcode.blockSignals(True)
            self.text_tokcode.clear()
            self.text_tokcode.blockSignals(False)
        if hasattr(self, "text_order"):
            self.text_order.blockSignals(True)
            self.text_order.clear()
            self.text_order.blockSignals(False)
        self.data_tokname.setText("")
        self.chk_uriagezero.setChecked(False)
        self.text_tokcode.setFocus()

    def update_settings(self):
        """変更ボタン (btn_exe_change) 処理"""
        code = self.text_tokcode.toPlainText().strip()
        new_sort_input = self.text_order.toPlainText().strip()

        code = "".join(filter(str.isdigit, code))
        new_sort_input = "".join(filter(str.isdigit, new_sort_input))

        if not code or not new_sort_input.isdigit():
            QMessageBox.critical(self, "エラー", "入力内容を確認してください。")
            return

        new_sort = int(new_sort_input)
        new_flag = 0 if self.chk_uriagezero.isChecked() else 1
        
        conn = common_utils.get_db_connection()
        cursor = conn.cursor()

        # try...finally による安全な切断管理
        try:
            cursor.execute("SELECT ISNULL(MAX(sort), 0) FROM Table_2 WHERE sort < 9999")
            max_sort_res = cursor.fetchone()
            max_sort = max_sort_res[0] if max_sort_res else 0

            if new_sort != 9999 and new_sort > max_sort:
                QMessageBox.critical(self, "入力エラー", f"表示順が大きすぎます。\n 最大値は {max_sort} です。")
                return

            cursor.execute("SELECT sort FROM Table_2 WHERE code = ?", (code,))
            old_sort_res = cursor.fetchone()
            if not old_sort_res:
                return

            old_sort = old_sort_res[0]
            if old_sort != new_sort:
                if old_sort < 9999 and new_sort < 9999:
                    if old_sort < new_sort:
                        cursor.execute("UPDATE Table_2 SET sort = sort - 1 WHERE sort > ? AND sort <= ? AND sort < 9999", (old_sort, new_sort))
                    else:
                        cursor.execute("UPDATE Table_2 SET sort = sort + 1 WHERE sort >= ? AND sort < ? AND sort < 9999", (new_sort, old_sort))
                elif old_sort == 9999 and new_sort <= max_sort:
                    cursor.execute("UPDATE Table_2 SET sort = sort + 1 WHERE sort >= ? AND sort < 9999", (new_sort,))
                elif old_sort < 9999 and new_sort == 9999:
                    cursor.execute("UPDATE Table_2 SET sort = sort - 1 WHERE sort > ? AND sort < 9999", (old_sort,))

            cursor.execute("UPDATE Table_2 SET sort = ?, flag = ? WHERE code = ?", (new_sort, new_flag, code))
            conn.commit()
            QMessageBox.information(self, "完了", f"得意先コード: {code}\n 設定を更新しました。")
            self.on_enter()
            self.text_tokcode.setFocus()
        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "エラー", f"失敗しました: {e}")
        finally:
            cursor.close()
            conn.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = UrikakeWindow()
    window.show()
    sys.exit(app.exec())
